# openpyxl 모듈은 워크시트 객체와 셀에 접속할 수 있는 메소드 제공
import openpyxl
import warnings
warnings.simplefilter("ignore")

# 전체 엑셀 시트(워크북) 를 메모리에 로드
workbook = openpyxl.load_workbook('excel_read.xlsx')

#엑셀 워크북에 존재하는 모든 워크시트의 이름 가져오기
print("워크북내의 워크시트들 : ", workbook.sheetnames)
print("=" * 30)

#고객 이름의 워크시트 가져오기
customer = workbook['고객']
print("고객 워크시트 객체 : ", customer)

# 셀 객체 읽기
# 이름 혹은 행/열 위치를 기준으로 셀을 읽기
print("첫번째 셀 객체 : ", customer['A1'])
print("첫번째 셀 객체의 값 : ", customer['A1'].value)

# 셀 객체는 cell() 메소드를 사용하여 접속
# 3행 2열의 셀 정보 가져오기
print("다른 cell 객체정보:", customer.cell(row=3, column=2))
print("다른 cell 객체가 가진 값:", customer.cell(row=3, column=2).value)

print("이름:", customer['B3'].value, customer['C3'].value)
